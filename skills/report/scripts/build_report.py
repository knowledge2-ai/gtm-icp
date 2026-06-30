#!/usr/bin/env python3
"""Build an investor-grade Excel workbook from scored ICP artifacts (the `report` stage).

Reads across the per-account artifacts each pipeline stage leaves under an
artifact root (`score.json`, `classify.json`, `enrich.json`, `people.json`,
`personalize.json`) and assembles a styled, multi-sheet `.xlsx` an operator can
present to investors:

  * Overview   — KPI band + pipeline funnel (counts computed from artifacts)
  * Market Size — TAM / SAM / SOM in dollars + addressable market by vertical
                  (only when --universe/--sample are supplied; see notes)
  * Methodology — each track's ICP rubric: gates, weighted dimensions, thresholds
  * <Track> Accounts — one sheet per track: firmographics, tier/score, the
                  per-dimension point breakdown, top contact, and rationale
  * Scoring Detail — the evidence behind every A-tier dimension score
  * Contacts   — every resolved contact, with cross-domain matches flagged

Multiple ICP tracks (e.g. Revenue + Design Partner) are passed as repeated
--track specs, each `ROOT:CRITERIA:LABEL` — the same companies scored against two
rubrics in two artifact roots. Dimension columns are derived from each criteria
file, so this is not pinned to any particular rubric.

The funnel counts (enriched, scored, A-tier, contacts, drafts) are all computed
from the artifacts. The addressable-market counts (universe / sample / per
vertical) are external Apollo figures the workbook cannot derive, so they are
explicit inputs — gather them with `apollo_search.py --count` and pass them in.
ACV is an assumption (--acv); every dollar figure rescales from it.

Requires openpyxl. stdlib otherwise; no network, no keys.
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.label import DataLabelList
except ImportError:
    sys.stderr.write(
        "The report skill needs openpyxl. Install it in your environment:\n"
        "    python3 -m pip install openpyxl\n")
    sys.exit(2)

# ---------- palette / shared styles ----------
NAVY = "1F3864"; LIGHT = "D9E1F2"; A_FILL = "C6EFCE"; B_FILL = "FFF2CC"
GREYF = "F2F2F2"; RED = "FFC7CE"; WHITE = "FFFFFF"
hdr_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
title_font = Font(name="Calibri", bold=True, color=NAVY, size=22)
sub_font = Font(name="Calibri", color="595959", size=11)
kpi_font = Font(name="Calibri", bold=True, color=NAVY, size=26)
kpilab = Font(name="Calibri", color="595959", size=10)
bold = Font(name="Calibri", bold=True, size=11)
small = Font(name="Calibri", size=10)
tiny = Font(name="Calibri", color="595959", size=9)
navy_fill = PatternFill("solid", fgColor=NAVY)
light_fill = PatternFill("solid", fgColor=LIGHT)
grey_fill = PatternFill("solid", fgColor=GREYF)
a_fill = PatternFill("solid", fgColor=A_FILL)
b_fill = PatternFill("solid", fgColor=B_FILL)
red_fill = PatternFill("solid", fgColor=RED)
_thin = Side(style="thin", color="BFBFBF")
border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center")
vtop = Alignment(vertical="top")


def _read(path: Path) -> dict:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def _revenue(e: dict):
    rev = e.get("annual_revenue")
    if isinstance(rev, (int, float)) and rev:
        return f"${rev / 1e6:.1f}M"
    return "n/a"


def load_accounts(root: Path) -> list[dict]:
    """One record per non-Reject scored account dir under `root`."""
    accts = []
    for d in sorted(p for p in root.iterdir() if p.is_dir() and not p.name.startswith("_")):
        sc = _read(d / "score.json")
        if not sc or sc.get("tier") == "Reject":
            continue
        e = _read(d / "enrich.json")
        cl = _read(d / "classify.json")
        dims = {x["key"]: x for x in cl.get("dimensions", []) if isinstance(x, dict)}
        ppl = [p for p in _read(d / "people.json").get("people", []) if isinstance(p, dict)]
        dom = (e.get("domain") or e.get("website") or "").lower().replace("www.", "")
        for p in ppl:
            em = (p.get("email") or "").lower()
            p["_mismatch"] = bool(em and dom and em.split("@")[-1] != dom)
        samedom = [p for p in ppl if p.get("email") and not p["_mismatch"]]
        withemail = [p for p in ppl if p.get("email")]
        top = (samedom or withemail or ppl or [{}])[0]
        accts.append({
            "slug": d.name,
            "company": sc.get("company_name") or e.get("company_name") or d.name,
            "domain": dom, "vertical": e.get("vertical") or e.get("industry") or "",
            "emp": e.get("employee_count") or "", "founded": e.get("founded_year") or "",
            "revenue": _revenue(e), "country": e.get("country") or "",
            "tier": sc.get("tier"), "score": sc.get("score"),
            "dims": dims, "rflags": [x for x in cl.get("red_flags", []) if x.get("present")],
            "people": ppl, "rationale": sc.get("rationale", ""),
            "top_name": top.get("name", ""), "top_title": top.get("title", ""),
            "top_email": top.get("email", ""), "top_li": top.get("linkedin_url", ""),
        })
    accts.sort(key=lambda a: -(a["score"] or 0))
    return accts


def count_drafts(root: Path) -> int:
    n = 0
    for d in root.iterdir():
        if not d.is_dir() or d.name.startswith("_"):
            continue
        pj = _read(d / "personalize.json")
        n += len(pj.get("drafts", []) or pj.get("messages", []))
    return n


def parse_track(spec: str) -> dict:
    parts = spec.split(":")
    root = Path(parts[0])
    criteria_path = Path(parts[1]) if len(parts) > 1 and parts[1] else None
    label = ":".join(parts[2:]) if len(parts) > 2 else root.name
    criteria = _read(criteria_path) if criteria_path else {}
    dim_keys = [d["key"] for d in criteria.get("dimensions", []) if isinstance(d, dict)]
    return {"root": root, "criteria": criteria, "label": label,
            "dim_keys": dim_keys, "accounts": load_accounts(root)}


# ---------- sheet builders ----------
def sheet_overview(wb, tracks, funnel, kpis, sources):
    ws = wb.active
    ws.title = "Overview"
    ws.sheet_view.showGridLines = False
    ws["A1"] = TITLE; ws["A1"].font = title_font
    ws["A2"] = SUBTITLE; ws["A2"].font = sub_font
    ws.merge_cells("A1:H1"); ws.merge_cells("A2:H2")
    for i, (val, lab, sub) in enumerate(kpis):
        col = 1 + (i % 3) * 3; row = 4 + (i // 3) * 4
        ws.cell(row=row, column=col, value=val).font = kpi_font
        ws.cell(row=row + 1, column=col, value=lab).font = bold
        ws.cell(row=row + 2, column=col, value=sub).font = kpilab
        for rr in range(row, row + 3):
            ws.cell(row=rr, column=col).fill = grey_fill
            ws.merge_cells(start_row=rr, start_column=col, end_row=rr, end_column=col + 1)
    fr = 13
    ws.cell(row=fr, column=1, value="Pipeline funnel").font = Font(bold=True, size=14, color=NAVY)
    ws.cell(row=fr + 1, column=1, value="Stage").font = hdr_font
    ws.cell(row=fr + 1, column=1).fill = navy_fill
    ws.cell(row=fr + 1, column=2, value="Count").font = hdr_font
    ws.cell(row=fr + 1, column=2).fill = navy_fill
    ws.cell(row=fr + 1, column=2).alignment = center
    for i, (stage, cnt) in enumerate(funnel):
        rr = fr + 2 + i
        ws.cell(row=rr, column=1, value=stage).font = small
        ws.cell(row=rr, column=2, value=cnt).font = bold
        ws.cell(row=rr, column=2).alignment = center
        fill = light_fill if i % 2 else grey_fill
        ws.cell(row=rr, column=1).fill = fill; ws.cell(row=rr, column=2).fill = fill
    nr = fr + 2 + len(funnel) + 1
    ws.cell(row=nr, column=1, value="Data sources").font = Font(bold=True, size=12, color=NAVY)
    ws.cell(row=nr + 1, column=1, value=sources).font = tiny
    ws.merge_cells(start_row=nr + 1, start_column=1, end_row=nr + 7, end_column=8)
    ws.cell(row=nr + 1, column=1).alignment = wrap
    ws.column_dimensions["A"].width = 50
    for c in "BCDEFGH":
        ws.column_dimensions[c].width = 14


def sheet_market(wb, mkt):
    ws = wb.create_sheet("Market Size")
    ws.sheet_view.showGridLines = False
    acv = mkt["acv"]; uni = mkt["universe"]; sam = mkt["sample"]
    hit = mkt["hit"]; projA = mkt["projA_sam"]
    tam_d = uni * acv; sam_d = sam * acv; som_d = projA * acv
    ws["A1"] = "Market Size & Opportunity"; ws["A1"].font = title_font
    ws["A2"] = (f"Bottom-up model · ACV assumption ${acv:,}/yr · company counts are live "
                f"Apollo match totals"); ws["A2"].font = sub_font
    ws.merge_cells("A1:G1"); ws.merge_cells("A2:G2")
    tiles = [("TAM — total addressable", f"${tam_d / 1e9:.1f}B", f"{uni:,} companies × ${acv // 1000}k ACV"),
             ("SAM — serviceable (priority verticals)", f"${sam_d / 1e9:.2f}B", f"{sam:,} companies × ${acv // 1000}k ACV"),
             ("SOM — near-term, high-fit A-tier", f"${som_d / 1e6:.0f}M", f"~{projA:,} A-tier × ${acv // 1000}k ACV")]
    for i, (lab, val, sub) in enumerate(tiles):
        col = 1 + i * 2
        ws.cell(row=4, column=col, value=val).font = kpi_font
        ws.cell(row=5, column=col, value=lab).font = bold
        ws.cell(row=6, column=col, value=sub).font = kpilab
        for rr in range(4, 7):
            ws.cell(row=rr, column=col).fill = grey_fill
            ws.merge_cells(start_row=rr, start_column=col, end_row=rr, end_column=col + 1)
    pr = 9
    ws.cell(row=pr, column=1, value="A-tier target pool").font = Font(bold=True, size=14, color=NAVY)
    pool = [("Validated this pass — individually scored", str(mkt["validated"]), mkt["validated_note"]),
            ("Observed A-tier hit rate", f"{hit * 100:.0f}%", f"{mkt['atier_n']} of {mkt['enriched']} enriched reached A-tier on ≥1 track"),
            ("Projected high-fit A-tier in SAM", f"~{projA:,}", f"{hit * 100:.0f}% × {sam:,} priority-vertical companies (extrapolation)"),
            ("Value of that A-tier pool (SOM)", f"${som_d / 1e6:.0f}M", f"~{projA:,} × ${acv // 1000}k ACV")]
    for j, h in enumerate(["Stage", "Count", "Basis"], 1):
        c = ws.cell(row=pr + 1, column=j, value=h); c.font = hdr_font; c.fill = navy_fill
        if j == 2:
            c.alignment = center
    for i, (s, c, b) in enumerate(pool):
        rr = pr + 2 + i
        ws.cell(row=rr, column=1, value=s).font = small
        ws.cell(row=rr, column=2, value=c).font = bold; ws.cell(row=rr, column=2).alignment = center
        ws.cell(row=rr, column=3, value=b).font = small
        fill = light_fill if i % 2 else grey_fill
        for cc in (1, 2, 3):
            ws.cell(row=rr, column=cc).fill = fill
        ws.merge_cells(start_row=rr, start_column=3, end_row=rr, end_column=7)
    vr = pr + 8
    vc = mkt.get("vertical_counts") or {}
    if vc:
        ws.cell(row=vr, column=1, value="Addressable market by vertical (live Apollo counts)").font = Font(bold=True, size=14, color=NAVY)
        for i, h in enumerate(["Vertical", "Companies", "Core sample", f"Market value (${acv // 1000}k ACV)"], 1):
            c = ws.cell(row=vr + 1, column=i, value=h); c.font = hdr_font; c.fill = navy_fill; c.border = border
            c.alignment = center if i > 1 else Alignment(vertical="center")
        core = set(mkt.get("core_tags", []))
        ints = [(t, n) for t, n in vc.items() if isinstance(n, int)]
        for j, (tag, n) in enumerate(sorted(ints, key=lambda kv: -kv[1])):
            rr = vr + 2 + j
            disp = tag.upper() if tag.isupper() or len(tag) <= 3 else tag.title()
            ws.cell(row=rr, column=1, value=disp).font = small
            ws.cell(row=rr, column=2, value=n).font = small
            ws.cell(row=rr, column=2).number_format = "#,##0"; ws.cell(row=rr, column=2).alignment = center
            if tag in core:
                ws.cell(row=rr, column=3, value="● core").font = small
                ws.cell(row=rr, column=3).fill = a_fill
            ws.cell(row=rr, column=3).alignment = center
            val = n * acv
            ws.cell(row=rr, column=4, value=f"${val / 1e9:.2f}B" if val >= 1e9 else f"${val / 1e6:.0f}M").font = small
            ws.cell(row=rr, column=4).alignment = center
            for cc in range(1, 5):
                ws.cell(row=rr, column=cc).border = border
        cav = vr + 2 + len(ints) + 1
    else:
        cav = vr
    ws.cell(row=cav, column=1, value=mkt["caveat"]).font = tiny
    ws.merge_cells(start_row=cav, start_column=1, end_row=cav + 5, end_column=7)
    ws.cell(row=cav, column=1).alignment = wrap
    # chart data parked in hidden cols S/T
    dc = 19
    ws.cell(row=3, column=dc, value="Tier"); ws.cell(row=3, column=dc + 1, value="$B")
    for i, (lab, val) in enumerate([("TAM", tam_d / 1e9), ("SAM", sam_d / 1e9), ("SOM", som_d / 1e9)]):
        ws.cell(row=4 + i, column=dc, value=lab); ws.cell(row=4 + i, column=dc + 1, value=round(val, 2))
    for col in (dc, dc + 1):
        ws.column_dimensions[get_column_letter(col)].hidden = True
    chart = BarChart(); chart.type = "bar"; chart.title = "Market size — TAM / SAM / SOM ($B)"
    chart.style = 10; chart.legend = None; chart.height = 6.5; chart.width = 13
    chart.add_data(Reference(ws, min_col=dc + 1, min_row=3, max_row=6), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=dc, min_row=4, max_row=6))
    chart.dataLabels = DataLabelList(); chart.dataLabels.showVal = True
    ws.add_chart(chart, "I3")
    ws.column_dimensions["A"].width = 42
    for c in "BCDEFG":
        ws.column_dimensions[c].width = 16


def sheet_methodology(wb, tracks):
    ws = wb.create_sheet("Methodology")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "ICP Methodology"; ws["A1"].font = title_font; ws.merge_cells("A1:F1")
    r = 3
    for t in tracks:
        C = t["criteria"]
        ws.cell(row=r, column=1, value=f"{t['label']} ICP  (root {t['root']})").font = Font(bold=True, size=14, color=NAVY)
        ws.cell(row=r + 1, column=1, value=C.get("bottom_line") or C.get("description", "")).font = sub_font
        ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=6)
        ws.cell(row=r + 1, column=1).alignment = wrap
        r += 3
        ws.cell(row=r, column=1, value="Hard gates (fail any → Reject)").font = bold
        for g in C.get("gates", []):
            r += 1
            ws.cell(row=r, column=1, value="• " + (g.get("name") or g.get("key", ""))).font = small
            ws.cell(row=r, column=2, value=g.get("description", "")).font = small
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            ws.cell(row=r, column=2).alignment = wrap
        r += 2
        ws.cell(row=r, column=1, value="Scored dimensions (weights)").font = bold
        ws.cell(row=r, column=1).fill = light_fill
        for d in C.get("dimensions", []):
            r += 1
            ws.cell(row=r, column=1, value="• " + (d.get("name") or d.get("key", ""))).font = small
            ws.cell(row=r, column=2, value=f"{d.get('max_points', '')} pts").font = bold
            ws.cell(row=r, column=3, value=d.get("description", "")).font = small
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
            ws.cell(row=r, column=3).alignment = wrap
        r += 2
        ws.cell(row=r, column=1, value=f"Tiers: {C.get('thresholds', {})}").font = sub_font
        r += 3
    ws.column_dimensions["A"].width = 30
    for c in "BCDEF":
        ws.column_dimensions[c].width = 22


def sheet_accounts(wb, track):
    accts = track["accounts"]; dim_keys = track["dim_keys"]
    ws = wb.create_sheet(f"{track['label']} Accounts"[:31])
    ws.sheet_view.showGridLines = False
    cols = (["Rank", "Company", "Domain", "Vertical", "Emp", "Founded", "Revenue", "Country",
             "Tier", "Score"] + [d.replace("_", " ").title() for d in dim_keys] +
            ["Red flags", "Top contact", "Title", "Email", "LinkedIn", "Why (scoring rationale)"])
    ws["A1"] = f"{track['label']} Accounts"; ws["A1"].font = title_font; ws.merge_cells("A1:H1")
    hrow = 3
    for i, h in enumerate(cols, 1):
        c = ws.cell(row=hrow, column=i, value=h); c.font = hdr_font; c.fill = navy_fill
        c.alignment = center; c.border = border
    for j, a in enumerate(accts):
        rr = hrow + 1 + j
        row = [j + 1, a["company"], a["domain"], a["vertical"], a["emp"], a["founded"],
               a["revenue"], a["country"], a["tier"], a["score"]]
        for dk in dim_keys:
            dd = a["dims"].get(dk)
            row.append(f"{dd['points_awarded']}/{dd['max_points']}" if dd else "")
        row += [", ".join(x["key"] for x in a["rflags"]) or "—", a["top_name"] or "—",
                a["top_title"], a["top_email"], a["top_li"], a["rationale"]]
        ws.append(row)
        fill = a_fill if a["tier"] == "A" else b_fill
        for cc in range(1, len(cols) + 1):
            cell = ws.cell(row=rr, column=cc); cell.border = border; cell.font = small; cell.alignment = vtop
        for cc in (9, 10):
            ws.cell(row=rr, column=cc).fill = fill; ws.cell(row=rr, column=cc).font = bold
            ws.cell(row=rr, column=cc).alignment = center
        ws.cell(row=rr, column=len(cols)).alignment = wrap
    ws.freeze_panes = ws.cell(row=hrow + 1, column=3)
    widths = [5, 26, 22, 20, 6, 8, 9, 12, 6, 7] + [9] * len(dim_keys) + [16, 20, 22, 26, 30, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_scoring_detail(wb, tracks):
    ws = wb.create_sheet("Scoring Detail")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Scoring detail — evidence behind each A-tier score"; ws["A1"].font = title_font
    ws.merge_cells("A1:F1")
    for i, h in enumerate(["Track", "Company", "Tier / Score", "Dimension", "Points", "Evidence"], 1):
        c = ws.cell(row=3, column=i, value=h); c.font = hdr_font; c.fill = navy_fill
        c.alignment = center; c.border = border
    rr = 4
    for t in tracks:
        for a in [x for x in t["accounts"] if x["tier"] == "A"]:
            for dk in t["dim_keys"]:
                dd = a["dims"].get(dk)
                if not dd:
                    continue
                ws.cell(row=rr, column=1, value=t["label"]).font = small
                ws.cell(row=rr, column=2, value=a["company"]).font = small
                ws.cell(row=rr, column=3, value=f"{a['tier']} / {a['score']:.0f}").font = bold
                ws.cell(row=rr, column=4, value=dk.replace("_", " ")).font = small
                ws.cell(row=rr, column=5, value=f"{dd['points_awarded']}/{dd['max_points']}").font = small
                ws.cell(row=rr, column=5).alignment = center
                ws.cell(row=rr, column=6, value=dd.get("evidence", "")).font = small
                ws.cell(row=rr, column=6).alignment = wrap
                for cc in range(1, 7):
                    ws.cell(row=rr, column=cc).border = border
                rr += 1
    for i, w in enumerate([14, 24, 12, 20, 8, 80], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


def sheet_contacts(wb, tracks):
    ws = wb.create_sheet("Contacts")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Verified contacts — actionable accounts"; ws["A1"].font = title_font
    ws.merge_cells("A1:H1")
    for i, h in enumerate(["Company", "Tier", "Contact", "Title", "Persona", "Email", "Email status", "LinkedIn"], 1):
        c = ws.cell(row=3, column=i, value=h); c.font = hdr_font; c.fill = navy_fill
        c.alignment = center; c.border = border
    # union by slug — prefer the first track that has the account (people are domain-based)
    bymap = {}
    for t in reversed(tracks):
        for a in t["accounts"]:
            bymap[a["slug"]] = a
    rr = 4
    for a in sorted(bymap.values(), key=lambda x: -(x["score"] or 0)):
        for p in a["people"]:
            ws.cell(row=rr, column=1, value=a["company"]).font = small
            ws.cell(row=rr, column=2, value=a["tier"]).font = bold; ws.cell(row=rr, column=2).alignment = center
            ws.cell(row=rr, column=2).fill = a_fill if a["tier"] == "A" else b_fill
            ws.cell(row=rr, column=3, value=p.get("name", "")).font = small
            ws.cell(row=rr, column=4, value=p.get("title", "")).font = small
            ws.cell(row=rr, column=5, value=p.get("persona", "")).font = small
            ws.cell(row=rr, column=6, value=p.get("email", "")).font = small
            estat = "⚠ verify (domain mismatch)" if p.get("_mismatch") else p.get("email_status", "")
            ws.cell(row=rr, column=7, value=estat).font = small; ws.cell(row=rr, column=7).alignment = center
            if p.get("_mismatch"):
                ws.cell(row=rr, column=7).fill = red_fill
            ws.cell(row=rr, column=8, value=p.get("linkedin_url", "")).font = small
            for cc in range(1, 9):
                ws.cell(row=rr, column=cc).border = border
            rr += 1
    ws.freeze_panes = "A4"
    for i, w in enumerate([26, 6, 22, 28, 22, 30, 14, 34], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return rr - 4


def build(tracks, mkt, out_path):
    primary = tracks[0]["root"]
    enriched = sum(1 for d in primary.iterdir()
                   if d.is_dir() and not d.name.startswith("_") and (d / "enrich.json").exists())
    scored = sum(1 for d in primary.iterdir()
                 if d.is_dir() and not d.name.startswith("_") and (d / "score.json").exists())
    # union of distinct A-tier companies (across tracks) and total contacts (dedup by slug)
    atier_slugs, contact_by_slug = set(), {}
    for t in tracks:
        for a in t["accounts"]:
            if a["tier"] == "A":
                atier_slugs.add(a["slug"])
            contact_by_slug[a["slug"]] = max(contact_by_slug.get(a["slug"], 0), len(a["people"]))
    contacts = sum(contact_by_slug.values())
    drafts = sum(count_drafts(t["root"]) for t in tracks)
    validated = sum(1 for t in tracks for a in t["accounts"] if a["tier"] == "A")

    kpis = []
    if mkt:
        kpis += [(f"{mkt['universe']:,}", "Addressable universe", "ICP-matched companies (Apollo)"),
                 (f"{mkt['sample']:,}", "Priority sample", "core verticals")]
    kpis += [(str(scored), "Scored on ICP rubric(s)", f"from {enriched} enriched"),
             (str(validated), "A-tier targets", " · ".join(f"{t['label']}: {sum(1 for a in t['accounts'] if a['tier']=='A')}" for t in tracks)),
             (str(contacts), "Verified contacts", "named decision-makers"),
             (str(drafts), "Outreach drafts", "grounded, per-contact")]

    funnel = []
    if mkt:
        funnel += [("Addressable universe — ICP-matched incumbents (Apollo)", f"{mkt['universe']:,}"),
                   ("Priority sample — core verticals", f"{mkt['sample']:,}")]
    funnel += [("Enriched & deep-scanned this pass", str(enriched)),
               ("Classified & scored (gate-screen passed)", str(scored))]
    for t in tracks:
        a = sum(1 for x in t["accounts"] if x["tier"] == "A")
        funnel.append((f"{t['label']} — actionable", f"{len(t['accounts'])}  ({a} A-tier)"))
    funnel += [("Distinct A-tier accounts with contacts", f"{len(atier_slugs)} → {contacts} contacts"),
               ("Grounded outreach drafts generated", str(drafts))]

    if mkt:
        mkt.update({"enriched": enriched, "atier_n": len(atier_slugs),
                    "hit": (len(atier_slugs) / enriched) if enriched else 0.0,
                    "validated": validated,
                    "validated_note": " · ".join(f"{sum(1 for a in t['accounts'] if a['tier']=='A')} {t['label']}" for t in tracks)})
        mkt["projA_sam"] = round(mkt["hit"] * mkt["sample"] / 100) * 100
        projA_tam = round(mkt["hit"] * mkt["universe"] / 1000) * 1000
        mkt["caveat"] = (
            "Notes: verticals are Apollo keyword tags and overlap — a company can match several — so rows do NOT sum "
            f"to the TAM; the de-duplicated universe is {mkt['universe']:,}. Per-vertical market value is illustrative "
            f"(count × ${mkt['acv'] // 1000}k) and inherits that overlap. The A-tier projection applies the hit rate "
            "observed on a high-fit precision slice; broad-universe yield is likely lower, so projecting to the full TAM "
            f"(~{projA_tam:,} A-tier) is an upper bound. ACV is an assumption — change it and every figure rescales.")

    src = (f"Company counts are Apollo company-search match totals, reproducible via `apollo_search.py --count`. "
           "Contacts and firmographics: Apollo enrichment. Intent signals: live scan of each company's "
           "website / careers / GitHub. Scores: deterministic rubric per the criteria file(s) on the Methodology tab. "
           + (f"Addressable universe {mkt['universe']:,} / priority sample {mkt['sample']:,}." if mkt else ""))

    wb = Workbook()
    sheet_overview(wb, tracks, funnel, kpis, src)
    if mkt:
        sheet_market(wb, mkt)
        wb.move_sheet("Market Size", -(len(wb.sheetnames) - 2))
    sheet_methodology(wb, tracks)
    for t in tracks:
        sheet_accounts(wb, t)
    sheet_scoring_detail(wb, tracks)
    n_contacts = sheet_contacts(wb, tracks)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return {"out": str(out_path), "sheets": wb.sheetnames, "tracks": [t["label"] for t in tracks],
            "enriched": enriched, "scored": scored, "a_tier": len(atier_slugs),
            "contacts": n_contacts, "drafts": drafts}


TITLE = "ICP Target Pipeline"
SUBTITLE = "Account qualification report"


def main(argv=None):
    global TITLE, SUBTITLE
    ap = argparse.ArgumentParser(description="Build the investor-grade ICP report workbook.")
    ap.add_argument("--track", action="append", default=[], metavar="ROOT:CRITERIA:LABEL",
                    help="ICP track to include; repeatable. e.g. .gtm:icp.criteria.json:Revenue")
    ap.add_argument("--criteria", help="criteria file for the default single track (when no --track given)")
    ap.add_argument("--out", help="output .xlsx path (default <primary-root>/_report/investor_report.xlsx)")
    ap.add_argument("--title", help="workbook title")
    ap.add_argument("--subtitle", help="workbook subtitle")
    ap.add_argument("--acv", type=int, default=50000, help="annual contract value for market sizing (default 50000)")
    ap.add_argument("--universe", type=int, help="addressable universe count (Apollo --count); enables Market Size sheet")
    ap.add_argument("--sample", type=int, help="priority-sample count (Apollo --count)")
    ap.add_argument("--vertical-counts", help="JSON file mapping vertical tag -> Apollo count")
    ap.add_argument("--core-tags", nargs="*", default=[], help="vertical tags to flag as the core sample")
    args = ap.parse_args(argv)

    if args.title:
        TITLE = args.title
    if args.subtitle:
        SUBTITLE = args.subtitle

    specs = args.track or [f"{os.environ.get('GTM_ARTIFACT_ROOT', '.gtm')}:{args.criteria or ''}:Accounts"]
    tracks = [parse_track(s) for s in specs]
    if not any(t["accounts"] for t in tracks):
        print(json.dumps({"error": "no scored accounts found under the track root(s); run score first"}))
        return 1

    mkt = None
    if args.universe and args.sample:
        mkt = {"acv": args.acv, "universe": args.universe, "sample": args.sample,
               "core_tags": args.core_tags,
               "vertical_counts": _read(Path(args.vertical_counts)) if args.vertical_counts else {}}

    out = Path(args.out) if args.out else tracks[0]["root"] / "_report" / "investor_report.xlsx"
    print(json.dumps(build(tracks, mkt, out)))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
